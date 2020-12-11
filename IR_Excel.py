#импортируем необходимые библиотеки
from colorama import init #для стиля ветового
init()#для стиля цветового
from openpyxl import load_workbook

#обозначим название файла с дублями и создадим сам файл
filename_1 = "Double.txt"
file = open(filename_1, "w")  # СОЗДАЕМ ФАЙЛ
file.close()

#создаем пустой текстовый файл
def myFile(list,text):
    """Функция запишет дублирующие данные в текстовый файл"""
    #print(type(list))
    file = open(filename_1, "a")  # открываем файл для записи дублей
    file.write(text)
    for i in list:
        file.write(i+"\n")
    file.close()

#для удобства вводим название файла
filename = "Список.xlsx"

#вытянули данные с документа
active_excel = load_workbook(filename=filename,data_only=True)#data_only=True

#делаем или смотрим активный лист
active_sheet = active_excel.active

#проверочное условие для определения корректности формата файла
if active_sheet["A1"].value == "Владелец":
    if active_sheet["B1"].value == "Характеристика":
        if active_sheet["C1"].value == "Единица измерения":
            if active_sheet["D1"].value == "Вид номенклатуры":
                if active_sheet["E1"].value == "Штрихкод":
                    if active_sheet["F1"].value == "Тип штрихкода":
                        #print(Fore.GREEN)
                        print("ФАЙЛ ПРИНЯТ И СООТВЕТСТВУЕТ ФОРМАТУ!")
else: print("\nФАЙЛ НЕ СООТВЕТСТВУЕТ ФОРМАТУ!")

# запишем наименование 7, 8 колонки
active_sheet["G1"] = 'Пробел'
active_sheet["H1"] = 'Сцепить'

#нужно понять максимальный размер данных на листе
max_row = active_sheet.max_row
max_column = active_sheet.max_column

print("\nКОЛИЧЕСТВО СТРОК: " + str(max_row))
print("КОЛИЧЕСТВО КОЛОНОК: " + str(max_column))

# вносим данные
space = " " #пробел для сцепки

#делаем цикл по заполнению пробелом всего диапазона колонки
for i in range(2,(max_row + 1)):
    _= active_sheet.cell(column= max_column - 1, row=i, value=space)

#делаем цикл по заполнению СЦЕПИТЬ всего диапазона колонки в екселе
for i in range(2,(max_row + 1)):
    a = 'A' + str(i)
    b = 'B' + str(i)
    c = 'G' + str(i)

    d = '=' + a + '&' + c + '&' + b
    _= active_sheet.cell(column= max_column, row=i, value=d)

#делаем цикл по заполнению СЦЕПИТЬ всего диапазона колонки в нашем списке
spisok = []
for i in range(2,(max_row + 1)):
    a_s = active_sheet["A" + str(i)].value
    b_s = active_sheet["B" + str(i)].value
    g_s = active_sheet["G" + str(i)].value

    d_s = str(a_s) + str(g_s) + str(b_s)
    spisok.append(d_s)

#анализируем ШК
barcode = []
for i in range(2,(max_row + 1)):
    e = active_sheet["E" + str(i)].value
    barcode.append(e)

unique_barcode = []
double_barcode = []
for i in barcode:
    # если количество вхождений = 1 - элемент уникальный
    if barcode.count(i) == 1: unique_barcode.append(i)
    else: double_barcode.append(i)

if len(barcode) == len(unique_barcode): print("\nДУБЛИКАТЫ ШК ОТСУТСТВУЮТ!")

else:
    d_barcode = (len(barcode) - len(unique_barcode))
    print("\nЕСТЬ ДУБЛИ ШК! " + str(d_barcode) + ' Шт:')

    # отправляем данные в функцию записи дублей
    myFile(double_barcode, "ДУБЛИ ШТРИХКОДОВ: (проверить наименования базы и файла!)\n")

    for i in double_barcode:
        print(i)

#анализируем наименование
unique_spisok = []
double_spisok = []
for i in spisok:
    if spisok.count(i) == 1: unique_spisok.append(i)
    else: double_spisok.append(i)

if len(spisok) == len(unique_spisok): print("\nДУБЛИКАТЫ НАИМЕНОВАНИЙ ОТСУТСТВУЮТ!")
else:
    print("\nЕСТЬ ДУБЛИ НАИМЕНОВАНИЙ! " + str(len(double_spisok)) + ' Шт:')

    # отправляем данные в функцию записи дублей
    myFile(double_spisok, "\n\nДУБЛИ НАИМЕНОВАНИЙ: (проверить штрихкода базы и в файле!)\n")

    for i in double_spisok:
        print(i)

# сохраняем изменения
active_excel.save("Список.xlsx") #сохраняем все изменения

input()
#pyinstaller -F IR_Excel_2.py
#+++++++++++++++